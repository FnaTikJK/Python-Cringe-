import csv
import datetime
from builtins import input
import re
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import matplotlib.pyplot as plt
from jinja2 import Environment, FileSystemLoader
import pdfkit
from os import path


class Vacancy:
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, object_vacancy):
        self.name = object_vacancy['name']
        salary_from = (int)((float)("".join(object_vacancy['salary_from'].split())))
        salary_to = (int)((float)("".join(object_vacancy['salary_to'].split())))
        self.salary = (salary_from + salary_to) * self.currency_to_rub[object_vacancy['salary_currency']] // 2
        self.area_name = object_vacancy['area_name']
        self.published_at = datetime.datetime.strptime(object_vacancy['published_at'], '%Y-%m-%dT%H:%M:%S%z')


class DataSet:
    def __init__(self, file_name: str, vacancies_objects: list):
        self.file_name = file_name
        self.vacancies_objects = vacancies_objects

    def cleanhtml(self, raw_html):
        cleantext = re.sub(re.compile('<.*?>'), '', raw_html)
        return cleantext

    def csv_read(self):
        list_naming = []
        vacancies = []
        with open(self.file_name, encoding='utf-8-sig') as r_file:
            file_reader = csv.reader(r_file, delimiter=",")
            count = 0
            for row in file_reader:
                if count == 0:
                    count += 1
                    list_naming = row
                else:
                    if "" in row or len(row) != len(list_naming):
                        continue
                    vacancies.append(row)
        if len(list_naming) == 0:
            print('Пустой файл')
            exit()
        if len(vacancies) == 0:
            print('Нет данных')
            exit()
        return (vacancies, list_naming)

    def csv_filer(self, reader, list_naming):
        vacancies = list()
        for row in reader:
            current = {}
            for i in range(len(row)):
                current[list_naming[i]] = row[i]
            vacancies.append(Vacancy(current))
        return vacancies

    def fill_vacancies(self):
        (vacancies, list_naiming) = self.csv_read()
        self.vacancies_objects = self.csv_filer(vacancies, list_naiming)


class CustomTuple:
    totalSalary = 0
    count = 0
    def __init__(self, totalSalary: int, count: int):
        self.totalSalary = totalSalary
        self.count = count


class InputConnect:
    years_stats = {
    }

    cities_stats = {
    }

    vacancy_stats = {
    }

    def start_input(self):
        self.file_name = input('Введите название файла: ')
        self.profession = input('Введите название профессии: ')
        self.city_count = 0

    def count_vacancies(self, vacancies: list):
        for vacancy in vacancies:
            self.city_count += 1
            year = int(vacancy.published_at.year)
            if year not in self.years_stats.keys():
                self.years_stats[year] = CustomTuple(vacancy.salary, 1)
                self.vacancy_stats[year] = CustomTuple(0, 0)
            else:
                self.years_stats[year].totalSalary += vacancy.salary
                self.years_stats[year].count += 1

            if vacancy.area_name not in self.cities_stats.keys():
                self.cities_stats[vacancy.area_name] = CustomTuple(vacancy.salary, 1)
            else:
                self.cities_stats[vacancy.area_name].totalSalary += vacancy.salary
                self.cities_stats[vacancy.area_name].count += 1

            if self.profession in vacancy.name:
                self.vacancy_stats[year].totalSalary += vacancy.salary
                self.vacancy_stats[year].count += 1

    def normalize_statistic(self):
        for year in self.years_stats.keys():
            self.years_stats[year].totalSalary = int(self.years_stats[year].totalSalary // self.years_stats[year].count)

        to_del = list()
        for city in self.cities_stats.keys():
            percent_count = round(self.cities_stats[city].count / self.city_count, 4)
            if percent_count < 0.01:
                to_del.append(city)
            else:
                self.cities_stats[city].totalSalary = int(self.cities_stats[city].totalSalary // self.cities_stats[city].count)
                self.cities_stats[city].count = percent_count
        for city in to_del:
            del [self.cities_stats[city]]

        for year in self.vacancy_stats.keys():
            if self.vacancy_stats[year].count != 0:
                self.vacancy_stats[year].totalSalary = int(self.vacancy_stats[year].totalSalary // self.vacancy_stats[year].count)

    def print_one(self, str_output: str, dict: dict, value_name: str):
        flag = False
        print(str_output, end='')
        ind = 0
        for year in dict.keys():
            if ind == 0:
                print(' {', end='')
                ind += 1
            printEnd = ', '
            if year == max(dict.keys()):
                printEnd = ''
                flag = True
            print(f"{year}: {getattr(dict[year], value_name)}", end=printEnd)
        if flag:
            print('}')

    def print_for_cities(self, str_output: str, dict: dict, names: list, value_name):
        flag = False
        print(str_output, end='')
        ind = 0
        for name in names:
            if ind == 0:
                print(' {', end='')
            printEnd = ', '
            if ind == len(names) - 1:
                printEnd = ''
                flag = True
            print(f"'{name}': {getattr(dict[name], value_name)}", end=printEnd)
            ind += 1
        if flag:
            print('}')

    def print_answer(self):
        self.print_one("Динамика уровня зарплат по годам:", self.years_stats, "totalSalary")
        self.print_one("Динамика количества вакансий по годам:", self.years_stats, "count")

        self.print_one("Динамика уровня зарплат по годам для выбранной профессии:", self.vacancy_stats, "totalSalary")
        self.print_one("Динамика количества вакансий по годам для выбранной профессии:", self.vacancy_stats, "count")

        cities_sorted = sorted(self.cities_stats, key=lambda x: self.cities_stats[x].totalSalary, reverse=True)
        del cities_sorted[10:]
        self.print_for_cities("Уровень зарплат по городам (в порядке убывания):", self.cities_stats,
                              cities_sorted, "totalSalary")
        cities_sorted = sorted(self.cities_stats, key=lambda x: self.cities_stats[x].count, reverse=True)
        del cities_sorted[10:]
        self.print_for_cities("Доля вакансий по городам (в порядке убывания):", self.cities_stats,
                              cities_sorted, "count")

    def get_sorted_cities(self, attr_name: str):
        current = {}
        sorted_names = sorted(self.cities_stats, key=lambda x: getattr(self.cities_stats[x], attr_name), reverse=True)
        del sorted_names[10:]
        for name in sorted_names:
            current[name] = self.cities_stats[name]
        return current


class Report:
    def generate_excel(self, inputer: InputConnect):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        wb.create_sheet("Статистика по годам")
        list = wb["Статистика по годам"]
        list['A1'] = "Год"
        list['A1'].font = Font(bold=True)
        list['A1'].border = thin_border
        list.column_dimensions['A'].width = 6
        list['B1'] = "Средняя зарплата"
        list['B1'].font = Font(bold=True)
        list['B1'].border = thin_border
        list.column_dimensions['B'].width = len("Средняя зарплата") + 2
        list['C1'] = f"Средняя зарплата - {inputer.profession}"
        list['C1'].font = Font(bold=True)
        list['C1'].border = thin_border
        list.column_dimensions['C'].width = len(f"Средняя зарплата - {inputer.profession}") + 2
        list['D1'] = "Количество вакансий"
        list['D1'].font = Font(bold=True)
        list['D1'].border = thin_border
        list.column_dimensions['D'].width = len("Количество вакансий") + 2
        list['E1'] = f"Количество вакансий - {inputer.profession}"
        list['E1'].font = Font(bold=True)
        list['E1'].border = thin_border
        list.column_dimensions['E'].width = len(f"Количество вакансий - {inputer.profession}") + 2
        for i in inputer.years_stats:
            list[f"A{i - 2005}"] = i
            list[f"A{i - 2005}"].border = thin_border
            list[f"B{i - 2005}"] = inputer.years_stats[i].totalSalary
            list[f"B{i - 2005}"].border = thin_border
            list[f"D{i - 2005}"] = inputer.years_stats[i].count
            list[f"D{i - 2005}"].border = thin_border
        for i in inputer.vacancy_stats:
            list[f"C{i - 2005}"] = inputer.vacancy_stats[i].totalSalary
            list[f"C{i - 2005}"].border = thin_border
            list[f"E{i - 2005}"] = inputer.vacancy_stats[i].count
            list[f"E{i - 2005}"].border = thin_border

        wb.create_sheet("Статистика по городам")
        list = wb["Статистика по городам"]
        list['A1'] = "Город"
        list['A1'].font = Font(bold=True)
        list['A1'].border = thin_border
        list.column_dimensions['A'].width = len("Город") + 2
        list['B1'] = "Уровень зарплат"
        list['B1'].font = Font(bold=True)
        list['B1'].border = thin_border
        list.column_dimensions['B'].width = len("Уровень зарплат") + 2
        list['D1'] = "Город"
        list['D1'].font = Font(bold=True)
        list['D1'].border = thin_border
        list.column_dimensions['D'].width = len("Город") + 2
        list['E1'] = "Доля ваканский"
        list['E1'].font = Font(bold=True)
        list['E1'].border = thin_border
        list.column_dimensions['E'].width = len("Доля ваканский") + 2
        sorted_cities = inputer.get_sorted_cities("totalSalary")
        ind = 2
        for i in sorted_cities:
            list[f"A{ind}"] = i
            list[f"A{ind}"].border = thin_border
            list.column_dimensions['A'].width = max(list.column_dimensions['A'].width, len(i) + 2)
            list[f"B{ind}"] = sorted_cities[i].totalSalary
            list[f"B{ind}"].border = thin_border
            ind = ind + 1
        sorted_cities = inputer.get_sorted_cities("count")
        ind = 2
        for i in sorted_cities:
            list[f"D{ind}"] = i
            list[f"D{ind}"].border = thin_border
            list.column_dimensions['D'].width = max(list.column_dimensions['D'].width, len(i) + 2)
            list[f"E{ind}"] = f"{round(sorted_cities[i].count * 100, 2)}%"
            list[f"E{ind}"].number_format = '0.00%'
            list[f"E{ind}"].border = thin_border
            ind = ind + 1
        wb.save("report.xlsx")

    def generate_image(self, inputer: InputConnect, filename: str):
        fig, axis = plt.subplots(2, 2)
        plt.rcParams.update({'font.size': 8})
        self.add_simple_graph(axis[0, 0], inputer.years_stats.keys(),
                              [inputer.years_stats[key].totalSalary for key in inputer.years_stats],
                              [inputer.vacancy_stats[key].totalSalary for key in inputer.vacancy_stats],
                        "Средняя з/п", f"з/п {inputer.profession}", "Уровень зарплат по годам")
        self.add_simple_graph(axis[0, 1], inputer.years_stats.keys(),
                              [inputer.years_stats[key].count for key in inputer.years_stats],
                              [inputer.vacancy_stats[key].count for key in inputer.vacancy_stats],
                        "Количество ваканский", f"Количество ваканский {inputer.profession}", "Количество вакансий по годам")
        sorted_cities_by_salary =  inputer.get_sorted_cities("totalSalary")
        self.add_horizontal_graph(axis[1, 0], [key for key in sorted_cities_by_salary],
                                  [sorted_cities_by_salary[key].totalSalary for key in sorted_cities_by_salary],
                            "Уровень зарплат по городам")
        sorted_cities_by_count = inputer.get_sorted_cities("count")
        self.add_circle_diagramm(axis[1, 1], [key for key in sorted_cities_by_salary],
                                 [sorted_cities_by_salary[key].count for key in sorted_cities_by_salary],
                             "Доля ваканский по городам")
        fig.set_size_inches(16, 9)
        fig.tight_layout(h_pad=1)
        plt.savefig(filename)

    def add_simple_graph(self, axis, x_val, y_val1, y_val2, name_values1, name_values2, title):
        axis.set_title(title, fontsize=16)
        axis.grid(axis="y")
        axis.bar([v + 0.2 for v in x_val], y_val2, label=name_values2, width=0.4)
        axis.bar([v - 0.2 for v in x_val], y_val1, label=name_values1, width=0.4)
        axis.legend()
        axis.tick_params(axis="x", labelrotation=90)

    def add_horizontal_graph(self, axis, x_val, y_val, title):
        axis.set_title(title, fontsize=16)
        axis.grid(axis="x")
        axis.barh(x_val, y_val)
        axis.invert_yaxis()

    def add_circle_diagramm(self, axis, names: list, values: list, title):
        axis.set_title(title, fontsize=16)
        names.append("Другие")
        values.append(1 - sum(values))
        axis.pie(values, labels=names)
        plt.axis('equal')

    def generate_pdf(self, inputer, template_name: str, filename: str):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template(template_name)
        render_rules = self.get_render_rules(inputer)
        pdf_template = template.render(render_rules)
        config = pdfkit.configuration(wkhtmltopdf=r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, filename, configuration=config, options={"enable-local-file-access": True})

    def get_render_rules(self, inputer):
        rules = {
            'profession': inputer.profession,
            'image_path': path.abspath("graph.png"),
        }
        for i in inputer.years_stats:
            rules[f"all_avg_salary{i}"] = inputer.years_stats[i].totalSalary
            rules[f"all_count{i}"] = inputer.years_stats[i].count
        for i in inputer.vacancy_stats:
            rules[f"profession_avg_salary{i}"] = inputer.vacancy_stats[i].totalSalary
            rules[f"profession_count{i}"] = inputer.vacancy_stats[i].count
        cities = inputer.get_sorted_cities("totalSalary")
        ind = 1
        for i in cities:
            rules[f"table1_city{ind}"] = i
            rules[f"city_salary{ind}"] = cities[i].totalSalary
            ind = ind + 1
        cities = inputer.get_sorted_cities("count")
        ind = 1
        for i in cities:
            rules[f"table2_city{ind}"] = i
            rules[f"city_count{ind}"] = f"{round(cities[i].count * 100, 2)}%"
            ind = ind + 1
        return rules


inputer = InputConnect()
inputer.start_input()
dataset = DataSet(inputer.file_name, list())
dataset.fill_vacancies()
inputer.count_vacancies(dataset.vacancies_objects)
inputer.normalize_statistic()
inputer.print_answer()

report = Report()
report.generate_excel(inputer)
report.generate_image(inputer, "graph.png")
report.generate_pdf(inputer, "pdf_template.html", "report.pdf")
